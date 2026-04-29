"""
bomc_xlsx_reviewer.py — Excel SBOM Review for LXCE BoMC (Bootable Media Creator)

Inherits from XlsxReviewer and overrides BoMC-specific phases.

BoMC template differences from UX:
  - No OUTPUT section — BoMC has no Target/... distribution path entries
  - No INPUT_ONECLI / INPUT_ELECTRON markers
  - All OSS entries are flat (groups: OSS, NPM, TRANSITIVE)
  - Header at openpyxl row 20 (0-indexed row 19), metadata rows 1-19

Phase map:
  0  IGNORE fill             — inherited
  1  Group classification     — overridden (OSS / NPM / TRANSITIVE / SKIP)
  2  Bidirectional match      — skipped (no OUTPUT group)
  3  OneCLI Col C placeholder — skipped (no INPUT_ONECLI group)
  4  Col I/J/K/L fill rules   — overridden (uses OSS group name)
  5  npm cross-reference      — inherited (package-lock.json only)
  6  License consistency      — inherited
  7  Col F/G validation       — inherited
  8  OneCLI SBOM xref         — skipped (no OneCLI SBOM for BoMC)
  9  Copyleft attention       — new (xorriso, cygwin, ffmpeg flags)
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Optional

from openpyxl.styles import PatternFill

from .xlsx_reviewer import (
    XlsxReviewer,
    ReviewResult,
    ORANGE, YELLOW, GRAY15, WHITE,
    COL_A, COL_C, COL_K,
    DEEP_DEP_MARKER,
    _cv,
    auto_detect_platform,
)

# ---------------------------------------------------------------------------
# BoMC-specific copyleft knowledge base
# ---------------------------------------------------------------------------

# Entries where Col C license string is imprecise → ORANGE on Col C
# Key: exact Col A value, Value: correct SPDX expression
_COL_C_IMPRECISE: dict[str, str] = {
    "ffmpeg.dll":    "LGPL-2.1-or-later",  # Windows: SBOM says LGPL-2.1
    "libffmpeg.so":  "LGPL-2.1-or-later",  # Linux:   SBOM says LGPL-2.1+ (deprecated SPDX+ notation)
}

# Electron sub-files (Col C = "see below") inherit their parent entry's Col K.
# Blank Col K on sub-files is correct — do NOT flag them.
_COL_K_NEEDS_YES: frozenset[str] = frozenset()

# Informational copyleft notes (no color change — already compliant)
_COPYLEFT_NOTES: dict[str, str] = {
    "xorriso": (
        "GPL-3.0-or-later — invoked via subprocess exec, not linked. "
        "GPL does not propagate to BoMC binary."
    ),
    "cygwin1.dll": (
        "LGPL-3.0-only + Cygwin Linking Exception — dynamically linked. "
        "Exception explicitly exempts dynamic linking; compliant."
    ),
    "cygiconv-2.dll": (
        "LGPL-3.0-only + Cygwin Linking Exception — same as cygwin1.dll; compliant."
    ),
}


# ---------------------------------------------------------------------------
# BomcXlsxReviewer
# ---------------------------------------------------------------------------
class BomcXlsxReviewer(XlsxReviewer):
    """Excel SBOM reviewer for BoMC. Inherits from XlsxReviewer."""

    def __init__(
        self,
        sbom_path: str,
        platform: Optional[str] = None,
        lock_path: Optional[str] = None,
        output_path: Optional[str] = None,
    ):
        super().__init__(
            sbom_path=sbom_path,
            platform=platform,
            deps_info_path=None,
            lock_path=lock_path,
            fossa_json_path=None,
            onecli_json_path=None,   # no OneCLI for BoMC
            output_path=output_path,
        )
        self._p9_findings: list[str] = []

    # ------------------------------------------------------------------
    # Phase 1 (override) — BoMC group classification
    # Groups: OSS, NPM, TRANSITIVE, SKIP
    # ------------------------------------------------------------------
    def _phase1_classify(self, ws) -> tuple[dict, Optional[int]]:
        groups: dict[int, str] = {}
        deep_dep_row: Optional[int] = None

        for row in range(1, ws.max_row + 1):
            a = _cv(ws, row, COL_A)
            c = _cv(ws, row, COL_C)
            f = _cv(ws, row, 6)  # Col F

            # Skip metadata rows, header, and structural markers
            if row <= 20:
                groups[row] = "SKIP"
                continue
            if a in ("Direct dependencies",):
                groups[row] = "SKIP"
                continue
            if a == DEEP_DEP_MARKER:
                groups[row] = "SKIP"
                deep_dep_row = row
                continue
            if not a:
                groups[row] = "SKIP"
                continue
            if c == "NOT DISTRIBUTED":
                groups[row] = "SKIP"
                continue

            if deep_dep_row and row > deep_dep_row:
                groups[row] = "TRANSITIVE"
                continue

            if f.lower() in ("npm", "npm (direct)", "npm (transitive)"):
                groups[row] = "NPM"
                continue

            groups[row] = "OSS"

        gc = Counter(groups.values())
        print(f"\n[Phase 1] Groups: {dict(gc)}")
        if deep_dep_row:
            print(f"          Deep dependencies separator at row {deep_dep_row}")
        return groups, deep_dep_row

    # ------------------------------------------------------------------
    # Phase 2 (override) — no-op for BoMC
    # ------------------------------------------------------------------
    def _phase2_match(self, ws, groups: dict) -> tuple[list, list, dict]:
        print("\n[Phase 2] SKIP — BoMC has no Output/Input bidirectional matching")
        return [], [], {}

    # ------------------------------------------------------------------
    # Phase 3 (override) — no-op for BoMC
    # ------------------------------------------------------------------
    def _phase3_onecli_col_c(self, ws, groups: dict) -> int:
        print("\n[Phase 3] SKIP — BoMC has no INPUT_ONECLI group")
        return 0

    # ------------------------------------------------------------------
    # Phase 4 (override) — same logic, BoMC-specific group names
    # ------------------------------------------------------------------
    def _phase4_ijkl(self, ws, groups: dict) -> None:
        from .xlsx_reviewer import _classify_license, COL_I, COL_J, COL_L
        bomc_groups = {"OSS", "NPM", "TRANSITIVE"}
        filled = skipped = 0

        for row, grp in groups.items():
            if grp not in bomc_groups:
                continue
            c_val = _cv(ws, row, COL_C)
            d_val = _cv(ws, row, 4)  # Col D

            if "PROPRIETARY" in c_val.upper() and d_val.upper() == "LENOVO DEVELOPED":
                for col in (COL_I, COL_J, COL_K, COL_L):
                    ws.cell(row=row, column=col).value = "N/A"
                filled += 1
                continue

            lc = _classify_license(c_val)
            if lc == "GPL":
                ws.cell(row=row, column=COL_I).value = "Yes"
                ws.cell(row=row, column=COL_J).value = "Yes"
                ws.cell(row=row, column=COL_K).value = "Yes"
                ws.cell(row=row, column=COL_L).value = "Required"
                filled += 1
            elif lc == "PERMISSIVE":
                for col, val in [(COL_I, "NO"), (COL_J, "NO"), (COL_K, "YES"), (COL_L, "Not required")]:
                    if not _cv(ws, row, col):
                        ws.cell(row=row, column=col).value = val
                filled += 1
            else:
                skipped += 1

        print(f"\n[Phase 4] I/J/K/L: Filled={filled}, Skipped(unknown/see-below)={skipped}")

    # ------------------------------------------------------------------
    # Phase 8 (override) — no-op for BoMC
    # ------------------------------------------------------------------
    def _phase8_onecli_xref(self, ws, groups: dict) -> tuple[int, int, list]:
        print("\n[Phase 8] SKIP — BoMC has no OneCLI SBOM cross-reference")
        return 0, 0, []

    # ------------------------------------------------------------------
    # Phase 9 (new) — copyleft attention flags
    # Called via _extra_phases hook before wb.save()
    # ------------------------------------------------------------------
    def _phase9_copyleft(self, ws, groups: dict) -> list[str]:
        findings: list[str] = []
        orange_c = 0
        yellow_k = 0
        noted_names: set[str] = set()  # deduplicate notes for repeated entries

        for row, grp in groups.items():
            if grp in ("SKIP", "IGNORE"):
                continue
            a = _cv(ws, row, COL_A)

            # Flag imprecise Col C license → ORANGE only when value needs correction
            if a in _COL_C_IMPRECISE:
                c_val = _cv(ws, row, COL_C)
                correct = _COL_C_IMPRECISE[a]
                if c_val.lower() not in ("see below", "") and c_val != correct:
                    ws.cell(row=row, column=COL_C).fill = ORANGE
                    findings.append(
                        f"[ORANGE Col C] Row {row}: {a!r} — Col C={c_val!r}, "
                        f"human should correct to {correct!r}"
                    )
                    orange_c += 1

            # Flag missing in_notice → YELLOW on Col K
            if a in _COL_K_NEEDS_YES:
                k_val = _cv(ws, row, COL_K)
                if not k_val:
                    ws.cell(row=row, column=COL_K).fill = YELLOW
                    findings.append(
                        f"[YELLOW Col K] Row {row}: {a!r} — in_notice blank, "
                        f"covered in parent entry"
                    )
                    yellow_k += 1

            # Print informational notes — once per unique name
            if a in _COPYLEFT_NOTES and a not in noted_names:
                note = _COPYLEFT_NOTES[a]
                findings.append(f"[NOTE] {a!r} — {note}")
                noted_names.add(a)

        print(f"\n[Phase 9] Copyleft attention: ORANGE(Col C)={orange_c}, "
              f"YELLOW(Col K)={yellow_k}, notes={len(noted_names)}")
        return findings

    # ------------------------------------------------------------------
    # Extension hook — calls Phase 9 before wb.save()
    # ------------------------------------------------------------------
    def _extra_phases(self, wb, ws, groups: dict) -> None:
        self._p9_findings = self._phase9_copyleft(ws, groups)

    # ------------------------------------------------------------------
    # Summary (override) — BoMC-specific group labels
    # ------------------------------------------------------------------
    def _build_summary(self, ws, groups: dict) -> dict:
        from .xlsx_reviewer import _fill_count
        row_sets = {
            "OSS":        {r for r, g in groups.items() if g == "OSS"},
            "npm":        {r for r, g in groups.items() if g == "NPM"},
            "Transitive": {r for r, g in groups.items() if g == "TRANSITIVE"},
        }
        summary = {}
        for label, row_set in row_sets.items():
            summary[label] = {
                "BLUE":    _fill_count(ws, row_set, COL_A, "ADD8E6"),
                "ORANGE":  _fill_count(ws, row_set, COL_A, "FFA500"),
                "PINK_A":  _fill_count(ws, row_set, COL_A, "FFB6C1"),
                "YEL_A":   _fill_count(ws, row_set, COL_A, "FFFF00"),
                "GRAY_C":  _fill_count(ws, row_set, COL_C, "D9D9D9"),
            }
        return summary

    # ------------------------------------------------------------------
    # Print report (override) — add Phase 9 section
    # ------------------------------------------------------------------
    def _print_report(self, r: ReviewResult) -> None:
        print("\n" + "=" * 70)
        print(f"  BoMC SBOM REVIEW SUMMARY — {r.platform.upper()}")
        print("=" * 70)

        hdr = f"{'Group':<16} | {'BLUE':>5} | {'ORANGE':>6} | {'PINK(A)':>7} | {'YEL(A)':>6} | {'GRAY(C)':>7}"
        print(hdr)
        print("-" * len(hdr))
        for label in ("OSS", "npm", "Transitive"):
            s = r.summary_rows.get(label, {})
            print(f"{label:<16} | {s.get('BLUE',0):>5} | {s.get('ORANGE',0):>6} | "
                  f"{s.get('PINK_A',0):>7} | {s.get('YEL_A',0):>6} | {s.get('GRAY_C',0):>7}")

        if r.phase5_unverified:
            print(f"\n=== TO BE VERIFIED (Yellow Col A) — {r.phase5_yellow} ===")
            for grp, row, name in sorted(r.phase5_unverified):
                print(f"  [{grp}] Row {row}: {name!r}")

        if r.phase7_f_pink or r.phase7_g_pink:
            print(f"\n=== Col F/G MISSING — F_pink={r.phase7_f_pink}, G_pink={r.phase7_g_pink} ===")

        if self._p9_findings:
            print(f"\n=== COPYLEFT ATTENTION ({len(self._p9_findings)} items) ===")
            for finding in self._p9_findings:
                print(f"  {finding}")

        print(f"\n[Done] Saved → {r.output_path}")


# ---------------------------------------------------------------------------
# Convenience factory
# ---------------------------------------------------------------------------
def review_bomc(
    sbom_path: str,
    platform: Optional[str] = None,
    lock_path: Optional[str] = None,
    output_path: Optional[str] = None,
) -> ReviewResult:
    reviewer = BomcXlsxReviewer(
        sbom_path=sbom_path,
        platform=platform,
        lock_path=lock_path,
        output_path=output_path,
    )
    return reviewer.run()
