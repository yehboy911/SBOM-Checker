"""
notice_reviewer.py — Third Party Notice ↔ SBOM cross-checker

Parses a Notice TXT (Package Title blocks) and compares against SBOM Excel
to find gaps (SBOM entries missing from Notice) and orphans (Notice entries
missing from SBOM).

Supports: ux (UpdateXpress), bomc (BoMC)

Usage (module):
  from sbom_checker.notice_reviewer import review_notice
  review_notice(notice_path, sbom_path)

Usage (CLI):
  sbom-checker review-notice <notice.txt> --sbom <sbom.xlsx> [--product {ux,bomc}]
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Columns (1-indexed for openpyxl)
# ---------------------------------------------------------------------------
_A, _B, _C, _E, _F = 1, 2, 3, 5, 6

_SKIP_C_EXACT = frozenset({"see below", "ignore", "not distributed"})
_SKIP_C_STARTSWITH = ("proprietary", "separate compliance")
_SKIP_A  = frozenset({"direct dependencies", "deep dependencies"})

# ---------------------------------------------------------------------------
# Notice TXT parsing
# ---------------------------------------------------------------------------
_TITLE_RE = re.compile(r'^Package Title:\s+(.+?)\s*$')
_VER_RE   = re.compile(r'^(.+?)\s+\(([^)]+)\)\s*$')


@dataclass(frozen=True)
class NoticeEntry:
    name: str
    version: str
    raw: str


def parse_notice(txt_path: str) -> list[NoticeEntry]:
    """Extract all 'Package Title: <name> (<version>)' entries from a Notice TXT."""
    entries: list[NoticeEntry] = []
    with open(txt_path, encoding="utf-8", errors="replace") as f:
        for line in f:
            m = _TITLE_RE.match(line.rstrip())
            if not m:
                continue
            raw = m.group(1)
            vm = _VER_RE.match(raw)
            if vm:
                entries.append(NoticeEntry(name=vm.group(1).strip(), version=vm.group(2).strip(), raw=raw))
            else:
                entries.append(NoticeEntry(name=raw.strip(), version="", raw=raw))
    return entries


# ---------------------------------------------------------------------------
# SBOM Excel reading
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class SbomEntry:
    row: int
    name: str
    version: str
    license: str


def _cv(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _is_output(ws, row: int) -> bool:
    e = _cv(ws, row, _E)
    return e.startswith("Target/") or e.startswith("Target\\")


def _is_onecli(ws, row: int) -> bool:
    return "onecli" in _cv(ws, row, _E).lower()


def read_sbom_entries(sbom_path: str, product: str) -> list[SbomEntry]:
    """
    Return top-level third-party entries from an SBOM Excel.

    Exclusions (both products):
      - Rows 1-20 (header/metadata)
      - Empty Col A, structural markers
      - Col C = "see below" (sub-entries inherit parent's Notice entry)
      - Col C = "IGNORE" or "NOT DISTRIBUTED"
      - Col C empty (no license — not a third-party package row)

    UX-only exclusions:
      - OUTPUT group (Col E starts with "Target/") — distribution targets
      - INPUT_ONECLI group (Col E contains "onecli") — OneCLI has its own Notice
    """
    wb = openpyxl.load_workbook(sbom_path, read_only=True, data_only=True)
    ws = wb.active
    entries: list[SbomEntry] = []

    for row in range(1, ws.max_row + 1):
        if row <= 20:
            continue
        a = _cv(ws, row, _A)
        if not a or a.lower() in _SKIP_A or a.startswith("*BUILD OUTPUT"):
            continue
        c = _cv(ws, row, _C)
        cl = c.lower()
        if not c or cl in _SKIP_C_EXACT or any(cl.startswith(p) for p in _SKIP_C_STARTSWITH):
            continue
        if product == "ux":
            if _is_output(ws, row) or _is_onecli(ws, row):
                continue
        entries.append(SbomEntry(row=row, name=a, version=_cv(ws, row, _B), license=c))

    wb.close()
    return entries


# ---------------------------------------------------------------------------
# Normalization & matching
# ---------------------------------------------------------------------------
_EXT_RE     = re.compile(r'\.(dll|so|exe|lib|a|dylib|node)(\.\d[\d.]*)?$', re.IGNORECASE)
_VER_SUF_RE = re.compile(r'[-_]\d+(\.\d+)*$')


def _normalize(name: str) -> str:
    n = name.strip().lower()
    n = _EXT_RE.sub('', n)       # strip file extension (.dll, .so, .exe …)
    n = _VER_SUF_RE.sub('', n)   # strip trailing -1.2.3 version suffix
    return n


@dataclass
class MatchResult:
    matched:        list[tuple[SbomEntry, NoticeEntry]] = field(default_factory=list)
    sbom_gaps:      list[SbomEntry]                     = field(default_factory=list)
    notice_orphans: list[NoticeEntry]                   = field(default_factory=list)


def match_entries(sbom: list[SbomEntry], notice: list[NoticeEntry]) -> MatchResult:
    """
    Match SBOM entries against Notice entries by normalized name.
    Deduplicates: multiple SBOM rows with the same normalized name count as one.
    Notice entries with the same name (different versions) are each reported separately
    if unmatched.
    """
    # notice lookup: normalized → list[NoticeEntry]
    notice_map: dict[str, list[NoticeEntry]] = {}
    for ne in notice:
        notice_map.setdefault(_normalize(ne.name), []).append(ne)

    result = MatchResult()
    seen_sbom:   set[str] = set()
    matched_keys: set[str] = set()

    for se in sbom:
        sn = _normalize(se.name)
        if sn in seen_sbom:
            continue
        seen_sbom.add(sn)
        if sn in notice_map:
            result.matched.append((se, notice_map[sn][0]))
            matched_keys.add(sn)
        else:
            result.sbom_gaps.append(se)

    for nn, nes in notice_map.items():
        if nn not in matched_keys:
            result.notice_orphans.extend(nes)

    return result


# ---------------------------------------------------------------------------
# Product auto-detection
# ---------------------------------------------------------------------------
def _auto_product(sbom_path: str) -> str:
    return "bomc" if "bomc" in Path(sbom_path).name.lower() else "ux"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def review_notice(
    notice_path: str,
    sbom_path: str,
    product: Optional[str] = None,
) -> MatchResult:
    product = product or _auto_product(sbom_path)

    print(f"\n[Init] Product: {product.upper()}")
    print(f"       Notice:  {Path(notice_path).name}")
    print(f"       SBOM:    {Path(sbom_path).name}")

    notice_entries = parse_notice(notice_path)
    print(f"\n[Notice] {len(notice_entries)} Package Title entries parsed")

    sbom_entries = read_sbom_entries(sbom_path, product)
    print(f"[SBOM]   {len(sbom_entries)} component entries (top-level, non-sub)")

    result = match_entries(sbom_entries, notice_entries)

    _print_report(result, product)
    return result


def _print_report(result: MatchResult, product: str) -> None:
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"  NOTICE REVIEW REPORT — {product.upper()}")
    print(sep)
    print(f"  Matched         : {len(result.matched)}")
    print(f"  SBOM gaps       : {len(result.sbom_gaps)}   ← in SBOM, NOT found in Notice")
    print(f"  Notice orphans  : {len(result.notice_orphans)}   ← in Notice, NOT found in SBOM")
    print(sep)

    if result.sbom_gaps:
        print(f"\n[GAP] {len(result.sbom_gaps)} SBOM entries missing from Notice:")
        for se in result.sbom_gaps:
            lic = se.license[:35] + "…" if len(se.license) > 35 else se.license
            print(f"  row {se.row:4d}  {se.name:<42} ver={se.version or '—':<15} lic={lic}")

    if result.notice_orphans:
        print(f"\n[ORPHAN] {len(result.notice_orphans)} Notice entries not matched in SBOM:")
        for ne in result.notice_orphans:
            print(f"  {ne.name:<42} ver={ne.version or '—'}")

    if not result.sbom_gaps and not result.notice_orphans:
        print("\n[OK] All SBOM entries matched in Notice. No gaps or orphans.")
