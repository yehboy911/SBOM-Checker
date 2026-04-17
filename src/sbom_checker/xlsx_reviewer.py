"""
xlsx_reviewer.py — Excel SBOM Review for LXCE UpdateXpress (UX)

Methodology: ~/.claude/rules/OSC-compliance-sbom-excel.md
IGNORE criteria: ~/.claude/skills/learned/OSC Scanning - Exclude File List.pdf
FOSSA framework: ~/.claude/skills/learned/Lenovo's Software Analysis Framework.pdf

Phases:
  0  IGNORE Fill  — image/pdf/doc extensions → Col C = "IGNORE"
  1  Group Classification
  2  Bidirectional Matching (Output ↔ Input)
  3  Input (OneCLI) Col C → "see below" + Gray 15%
  4  Col I/J/K/L Fill Rules
  5  npm Cross-reference (Col A Yellow)
  6  License Consistency (Col C: "see below"→Gray, empty→White)
  7  Col F/G Validation
  8  OneCLI SBOM Cross-reference (onecli_*_sbom_data.json)
  9  Save & Summary

Usage (CLI):
  sbom-checker review-xlsx <xlsx_path> [--platform linux|windows]
                           [--deps-info PATH] [--lock PATH]
                           [--fossa-json PATH] [--onecli-json PATH]
                           [--output PATH]

Usage (module):
  from sbom_checker.xlsx_reviewer import XlsxReviewer, auto_detect_platform
  reviewer = XlsxReviewer.from_args(sbom_path, platform="linux", ...)
  result = reviewer.run()
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
BLUE   = PatternFill("solid", start_color="ADD8E6", end_color="ADD8E6")
PINK   = PatternFill("solid", start_color="FFB6C1", end_color="FFB6C1")
ORANGE = PatternFill("solid", start_color="FFA500", end_color="FFA500")
YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
GRAY15 = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
WHITE  = PatternFill("none")

# Columns (1-indexed for openpyxl)
COL_A, COL_B, COL_C, COL_D, COL_E, COL_F, COL_G = 1, 2, 3, 4, 5, 6, 7
COL_I, COL_J, COL_K, COL_L = 9, 10, 11, 12

DEEP_DEP_MARKER = "Deep dependencies"

# ---------------------------------------------------------------------------
# IGNORE extension set — FOSSA Review column ONLY
# Ref: ~/.claude/skills/learned/OSC Scanning - Exclude File List.pdf
#
# Category    | Extensions                          | FOSSA Review
# ------------|-------------------------------------|-------------
# Picture     | gif png bmp jpeg jpg svg ico wav    | X (IGNORE)
# Web pages   | html htm shtm                       | X (IGNORE)
# Text        | doc docx xls xlsx ppt txt cvs ttf   | X (IGNORE)
#             | odt pdf ChangeLog                   | X (IGNORE)
#
# NOT ignored by FOSSA: Audio/video, Script, Binaries (.so .dll .exe .bin .lib),
# Build files, Version management. Binaries appear in SBOM and must be reviewed.
# ---------------------------------------------------------------------------
IGNORE_EXTENSIONS: frozenset[str] = frozenset({
    # Picture
    ".gif", ".png", ".bmp", ".jpeg", ".jpg", ".svg", ".ico", ".wav",
    # Web pages
    ".html", ".htm", ".shtm",
    # Text / docs
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".txt", ".cvs",
    ".ttf", ".odt", ".pdf",
})

# ---------------------------------------------------------------------------
# File-extension → Col F (USE) mapping
# ---------------------------------------------------------------------------
EXT_RE = re.compile(r'\.[a-zA-Z0-9]{2,6}$')

_USE_MAP_COMMON = {
    ".dll": "DLL", ".exe": "EXE", ".pak": "PAK", ".zip": "ZIP",
    ".asar": "ASAR", ".dat": "File", ".sys": "File", ".cat": "File",
    ".txt": "File", ".pdf": "File", ".png": "File", ".gif": "File",
}
_USE_MAP_LINUX_EXTRA = {".so": "SO", ".tgz": "TGZ", ".rpm": "RPM"}


def _ext_to_use(filename: str, platform: str) -> Optional[str]:
    m = EXT_RE.search(filename.lower())
    if not m:
        return None
    ext = m.group()
    use_map = {**_USE_MAP_COMMON}
    if platform == "linux":
        use_map.update(_USE_MAP_LINUX_EXTRA)
    if ext == ".bin":
        name_lower = filename.lower()
        return "Binary" if ("snapshot" in name_lower or "blob" in name_lower) else "BIN"
    return use_map.get(ext, "File")


# ---------------------------------------------------------------------------
# Platform configuration
# ---------------------------------------------------------------------------
_PA_RAW = (
    "/Users/Yehboy/Library/CloudStorage/OneDrive-Lenovo/"
    "文件/OSC/LXCE/UpdateXpress/LXCE_UpdateXpress_v5.4.0/PA/RAW"
)

@dataclass
class PlatformConfig:
    name: str
    # Group classification path markers
    output_target_substr: str   # must appear in Col E for OUTPUT
    output_target_fallback: str  # fallback check (e.g. "lnvgy_utl_lxce_ux")
    input_onecli_marker: str
    input_electron_marker: str
    # OneCLI cross-reference JSON
    default_onecli_json: str
    # Known ORANGE exceptions (exact name or callable predicate)
    known_orange_names: frozenset[str] = field(default_factory=frozenset)
    # Distribution package predicate (returns True if name = top-level dist pkg)
    dist_pkg_predicate: object = None  # callable(name) -> bool


def _linux_dist_pkg(name: str) -> bool:
    return name.startswith("lnvgy_utl_lxce_ux") and name.endswith("_linux_indiv.tgz")


def _windows_dist_pkg(name: str) -> bool:
    return name.startswith("lnvgy_utl_lxce_ux") and (
        name.endswith("_windows_indiv.zip") or name.endswith("_windows_indiv (zip)")
    )


PLATFORM_CONFIGS: dict[str, PlatformConfig] = {
    "linux": PlatformConfig(
        name="linux",
        output_target_substr="Target/lnvgy_utl_lxce_ux_linux_x86-64/",
        output_target_fallback="lnvgy_utl_lxce_ux",
        input_onecli_marker="updatexpress/bin/command/onecli/",
        input_electron_marker="updatexpress/bin/electron/",
        default_onecli_json=f"{_PA_RAW}/onecli_linux_sbom_data.json",
        # lxce_ux.bin = Electron binary renamed (Linux); app.asar = ASAR bundle
        # icudtl.dat intentionally NOT here → triggers manual-action reminder
        known_orange_names=frozenset({"lxce_ux.bin", "app.asar"}),
        dist_pkg_predicate=_linux_dist_pkg,
    ),
    "windows": PlatformConfig(
        name="windows",
        output_target_substr="Target\\lnvgy_utl_lxce_ux_windows_x86-64\\",
        output_target_fallback="lnvgy_utl_lxce_ux",
        input_onecli_marker="updatexpress\\bin\\command\\onecli\\",
        input_electron_marker="updatexpress\\bin\\electron\\",
        default_onecli_json=f"{_PA_RAW}/onecli_win_sbom_data.json",
        # lxce_ux.exe = Electron binary renamed (Windows); app.asar = ASAR bundle
        # icudtl.dat intentionally NOT here → triggers manual-action reminder
        known_orange_names=frozenset({"lxce_ux.exe", "app.asar"}),
        dist_pkg_predicate=_windows_dist_pkg,
    ),
}


def auto_detect_platform(xlsx_path: str) -> str:
    """Detect platform from SBOM filename. Returns 'linux' or 'windows'."""
    name_lower = Path(xlsx_path).name.lower()
    if "linux" in name_lower or "(tgz)" in name_lower or "tgz" in name_lower:
        return "linux"
    if "windows" in name_lower or "(zip)" in name_lower or "zip" in name_lower:
        return "windows"
    raise ValueError(
        f"Cannot auto-detect platform from '{Path(xlsx_path).name}'. "
        "Pass --platform linux or --platform windows explicitly."
    )


# ---------------------------------------------------------------------------
# License classification
# ---------------------------------------------------------------------------
_PERMISSIVE_TOKENS = frozenset({
    "mit", "isc", "bsd", "bsd-2-clause", "bsd-3-clause",
    "apache", "apache-2.0", "x11", "unicode", "zlib", "cc0",
    "cc0-1.0", "wtfpl", "unlicense", "0bsd", "ms-eula",
    "commercial", "proprietary", "public domain",
})


def _classify_license(c_val: str) -> str:
    if not c_val or c_val.lower() in ("see below", "ignore"):
        return "UNKNOWN"
    c_lower = c_val.lower()
    if "lgpl" in c_lower or ("gpl" in c_lower and "lgpl" not in c_lower):
        return "GPL"
    tokens = [t.strip().lower() for t in c_lower.replace(";", ",").split(",") if t.strip()]
    if tokens and all(any(p in t for p in _PERMISSIVE_TOKENS) for t in tokens):
        return "PERMISSIVE"
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def _cv(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _fill_count(ws, row_set: set, col: int, hex_color: str) -> int:
    count = 0
    hex_upper = hex_color.upper()
    for row in row_set:
        cell = ws.cell(row=row, column=col)
        if cell.fill and cell.fill.fill_type == "solid":
            if hex_upper in cell.fill.fgColor.rgb.upper():
                count += 1
    return count


# ---------------------------------------------------------------------------
# Review result
# ---------------------------------------------------------------------------
@dataclass
class ReviewResult:
    platform: str
    sbom_path: str
    output_path: str
    group_counts: dict
    deep_dep_row: Optional[int]
    phase2_stats: dict
    orange_list: list[str]
    pink_list: list[str]
    phase8_blue: int
    phase8_orange: int
    phase8_orange_list: list[tuple]
    phase5_yellow: int
    phase5_unverified: list[tuple]
    phase7_f_pink: int
    phase7_g_pink: int
    ignore_filled: int
    manual_actions: list[str]
    summary_rows: dict


# ---------------------------------------------------------------------------
# Main reviewer
# ---------------------------------------------------------------------------
class XlsxReviewer:
    """
    Excel SBOM reviewer for LXCE UpdateXpress.

    Parameters
    ----------
    sbom_path : str
    platform : str              'linux' or 'windows'
    deps_info_path : str | None  dependencies_info.json (primary npm ref)
    lock_path : str | None       package-lock.json (fallback npm ref)
    fossa_json_path : str | None FOSSA export JSON (tertiary npm ref)
    onecli_json_path : str | None override default OneCLI SBOM JSON
    output_path : str | None     defaults to <original>_reviewed.xlsx
    """

    def __init__(
        self,
        sbom_path: str,
        platform: Optional[str] = None,
        deps_info_path: Optional[str] = None,
        lock_path: Optional[str] = None,
        fossa_json_path: Optional[str] = None,
        onecli_json_path: Optional[str] = None,
        output_path: Optional[str] = None,
    ):
        self.sbom_path = sbom_path
        self.platform = platform or auto_detect_platform(sbom_path)
        self.cfg = PLATFORM_CONFIGS[self.platform]
        self.deps_info_path = deps_info_path
        self.lock_path = lock_path
        self.fossa_json_path = fossa_json_path
        self.onecli_json_path = onecli_json_path or self.cfg.default_onecli_json
        self.output_path = output_path or sbom_path.replace("-DRAFT.xlsx", "-DRAFT_reviewed.xlsx")

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------
    def run(self) -> ReviewResult:
        print(f"\n[Init] Platform: {self.platform.upper()}  SBOM: {Path(self.sbom_path).name}")

        known_npm = self._load_npm_refs()
        wb = openpyxl.load_workbook(self.sbom_path)
        ws = wb.active
        print(f"[Load] {ws.max_row} rows, {ws.max_column} columns")

        groups, deep_dep_row = self._phase1_classify(ws)
        ignore_filled       = self._phase0_ignore(ws, groups)
        groups              = self._phase1_reclassify_ignore(ws, groups)
        orange_list, pink_list, p2_stats = self._phase2_match(ws, groups)
        p3_count            = self._phase3_onecli_col_c(ws, groups)
        self._phase4_ijkl(ws, groups)
        p5_yellow, p5_unverified = self._phase5_npm(ws, groups, known_npm)
        self._phase6_license_c(ws, groups)
        p7_f, p7_g          = self._phase7_fg_validate(ws, groups)
        p8_blue, p8_orange, p8_orange_list = self._phase8_onecli_xref(ws, groups)

        wb.save(self.output_path)
        print(f"\n[Save] → {self.output_path}")

        # Count fills for summary
        wb2 = openpyxl.load_workbook(self.output_path)
        ws2 = wb2.active
        summary_rows = self._build_summary(ws2, groups)

        manual_actions = self._detect_manual_actions(orange_list)

        result = ReviewResult(
            platform=self.platform,
            sbom_path=self.sbom_path,
            output_path=self.output_path,
            group_counts=dict(Counter(groups.values())),
            deep_dep_row=deep_dep_row,
            phase2_stats=p2_stats,
            orange_list=orange_list,
            pink_list=pink_list,
            phase8_blue=p8_blue,
            phase8_orange=p8_orange,
            phase8_orange_list=p8_orange_list,
            phase5_yellow=p5_yellow,
            phase5_unverified=p5_unverified,
            phase7_f_pink=p7_f,
            phase7_g_pink=p7_g,
            ignore_filled=ignore_filled,
            manual_actions=manual_actions,
            summary_rows=summary_rows,
        )
        self._print_report(result)
        return result

    # ------------------------------------------------------------------
    # npm reference loading (Phase 5 sources)
    # ------------------------------------------------------------------
    def _load_npm_refs(self) -> set[str]:
        known: set[str] = set()

        # Priority 1: dependencies_info.json
        if self.deps_info_path and Path(self.deps_info_path).exists():
            with open(self.deps_info_path) as f:
                dinfo = json.load(f)
            known.update(dinfo.keys())
            print(f"[Phase 5] dependencies_info.json: {len(dinfo)} names loaded")
        else:
            print("[Phase 5] WARN: dependencies_info.json not provided or not found")

        # Priority 2: package-lock.json
        if self.lock_path and Path(self.lock_path).exists():
            with open(self.lock_path) as f:
                lock = json.load(f)
            lock_names: set[str] = set()
            if "packages" in lock:
                for pkg_path in lock["packages"]:
                    if pkg_path and "node_modules/" in pkg_path:
                        lock_names.add(pkg_path.split("node_modules/")[-1])
            elif "dependencies" in lock:
                lock_names.update(lock["dependencies"].keys())
            new = lock_names - known
            known.update(new)
            print(f"[Phase 5] package-lock.json: +{len(new)} transitive names")
        else:
            print("[Phase 5] WARN: package-lock.json not provided or not found")

        # Priority 3: FOSSA JSON (directDependencies with source=npm)
        # Ref: ~/.claude/skills/learned/Lenovo's Software Analysis Framework.pdf
        if self.fossa_json_path and Path(self.fossa_json_path).exists():
            with open(self.fossa_json_path) as f:
                fossa = json.load(f)
            fossa_npm: set[str] = set()
            for item in fossa.get("directDependencies", []):
                if item.get("source") == "npm":
                    fossa_npm.add(item["package"])
            new = fossa_npm - known
            known.update(new)
            print(f"[Phase 5] FOSSA JSON (npm direct deps): +{len(new)} names  ({len(fossa_npm)} total npm in FOSSA)")
        else:
            print("[Phase 5] INFO: FOSSA JSON not provided (optional, use --fossa-json)")

        print(f"[Phase 5] Total known npm names: {len(known)}")
        return known

    # ------------------------------------------------------------------
    # Phase 0 — IGNORE Fill
    # Ref: ~/.claude/skills/learned/OSC Scanning - Exclude File List.pdf
    # Extensions: .gif .png .bmp .jpeg .jpg .svg .ico .pdf .wav .html .htm
    # ------------------------------------------------------------------
    def _phase0_ignore(self, ws, groups: dict) -> int:
        filled = 0
        for row, grp in groups.items():
            if grp == "SKIP":
                continue
            a = _cv(ws, row, COL_A)
            m = EXT_RE.search(a.lower())
            if m and m.group() in IGNORE_EXTENSIONS:
                c = _cv(ws, row, COL_C)
                if not c:
                    ws.cell(row=row, column=COL_C).value = "IGNORE"
                    filled += 1
        print(f"\n[Phase 0] IGNORE fill: {filled} rows (extensions: gif/png/bmp/jpg/jpeg/svg/ico/pdf/wav/html/htm)")
        return filled

    def _phase1_reclassify_ignore(self, ws, groups: dict) -> dict:
        """After Phase 0, rows with Col C='IGNORE' get group=IGNORE for downstream skipping."""
        for row, grp in groups.items():
            if grp == "SKIP":
                continue
            if _cv(ws, row, COL_C).upper() == "IGNORE":
                groups[row] = "IGNORE"
        return groups

    # ------------------------------------------------------------------
    # Phase 1 — Group Classification
    # ------------------------------------------------------------------
    def _phase1_classify(self, ws) -> tuple[dict, Optional[int]]:
        groups: dict[int, str] = {}
        deep_dep_row: Optional[int] = None
        cfg = self.cfg

        for row in range(1, ws.max_row + 1):
            a = _cv(ws, row, COL_A)
            c = _cv(ws, row, COL_C)
            e = _cv(ws, row, COL_E)
            f = _cv(ws, row, COL_F)

            if row <= 20 or a.startswith("*BUILD OUTPUT") or a in ("Direct dependencies",):
                groups[row] = "SKIP"
                continue
            if a == DEEP_DEP_MARKER:
                groups[row] = "SKIP"
                deep_dep_row = row
                continue
            if not a:
                groups[row] = "SKIP"
                continue

            if deep_dep_row and row > deep_dep_row:
                groups[row] = "TRANSITIVE"
                continue

            if cfg.output_target_substr in e:
                groups[row] = "OUTPUT"
                continue
            if e.startswith("Target/") and cfg.output_target_fallback in e:
                groups[row] = "OUTPUT"
                continue
            if e.startswith("Target\\") and cfg.output_target_fallback in e:
                groups[row] = "OUTPUT"
                continue
            if e.startswith("1/"):
                groups[row] = "INPUT_SOURCE"
                continue
            if cfg.input_onecli_marker in e:
                groups[row] = "INPUT_ONECLI"
                continue
            if cfg.input_electron_marker in e:
                groups[row] = "INPUT_ELECTRON"
                continue
            if f.lower() == "npm" or f.lower().startswith("npm"):
                groups[row] = "NPM"
                continue
            if c == "NOT DISTRIBUTED":
                groups[row] = "SKIP"
                continue

            groups[row] = "SKIP"

        gc = Counter(groups.values())
        print(f"\n[Phase 1] Groups: {dict(gc)}")
        if deep_dep_row:
            print(f"          Deep dependencies separator at row {deep_dep_row}")
        return groups, deep_dep_row

    # ------------------------------------------------------------------
    # Phase 2 — Bidirectional Matching
    # ------------------------------------------------------------------
    def _phase2_match(self, ws, groups: dict) -> tuple[list, list, dict]:
        output_names: dict[str, list[int]] = defaultdict(list)
        input_names:  dict[str, list[int]] = defaultdict(list)

        for row, grp in groups.items():
            a = _cv(ws, row, COL_A)
            if grp == "OUTPUT":
                output_names[a].append(row)
            elif grp in ("INPUT_SOURCE", "INPUT_ONECLI", "INPUT_ELECTRON"):
                input_names[a].append(row)

        print(f"\n[Phase 2] output_names: {len(output_names)} unique | input_names: {len(input_names)} unique")

        stats = {"BLUE_out": 0, "ORANGE_out": 0, "BLUE_in": 0, "PINK_in": 0, "SKIP_in": 0}
        orange_list: list[str] = []
        pink_list: list[str] = []

        # Pass A: Output → Input
        for name, rows in output_names.items():
            if name in input_names:
                for row in rows:
                    ws.cell(row=row, column=COL_A).fill = BLUE
                stats["BLUE_out"] += len(rows)
            else:
                for row in rows:
                    ws.cell(row=row, column=COL_A).fill = ORANGE
                stats["ORANGE_out"] += len(rows)
                orange_list.append(name)

        # Pass B: Input → Output
        for name, rows in input_names.items():
            skip_all = all(
                "PROPRIETARY" in _cv(ws, r, COL_C).upper()
                and _cv(ws, r, COL_D).upper() == "LENOVO DEVELOPED"
                for r in rows
            )
            if skip_all:
                stats["SKIP_in"] += len(rows)
                continue
            if name in output_names:
                for row in rows:
                    ws.cell(row=row, column=COL_A).fill = BLUE
                stats["BLUE_in"] += len(rows)
            else:
                for row in rows:
                    ws.cell(row=row, column=COL_A).fill = PINK
                    if not _cv(ws, row, COL_C):
                        ws.cell(row=row, column=COL_C).value = "NOT DISTRIBUTED"
                stats["PINK_in"] += len(rows)
                pink_list.append(name)

        print(f"[Phase 2] Output: BLUE={stats['BLUE_out']}, ORANGE={stats['ORANGE_out']}")
        print(f"[Phase 2] Input:  BLUE={stats['BLUE_in']}, PINK={stats['PINK_in']}, "
              f"Skipped(PROP+LENOVO)={stats['SKIP_in']}")
        return orange_list, pink_list, stats

    # ------------------------------------------------------------------
    # Phase 3 — Input (OneCLI) Col C placeholder
    # ------------------------------------------------------------------
    def _phase3_onecli_col_c(self, ws, groups: dict) -> int:
        filled = 0
        for row, grp in groups.items():
            if grp != "INPUT_ONECLI":
                continue
            if not _cv(ws, row, COL_C):
                ws.cell(row=row, column=COL_C).value = "see below"
                ws.cell(row=row, column=COL_C).fill = GRAY15
                filled += 1
        print(f"\n[Phase 3] Input(OneCLI) Col C → 'see below': {filled} rows")
        return filled

    # ------------------------------------------------------------------
    # Phase 4 — Col I/J/K/L Fill Rules
    # ------------------------------------------------------------------
    def _phase4_ijkl(self, ws, groups: dict) -> None:
        input_groups = {"INPUT_SOURCE", "INPUT_ONECLI", "INPUT_ELECTRON", "NPM", "TRANSITIVE"}
        filled = skipped = 0

        for row, grp in groups.items():
            if grp not in input_groups:
                continue
            c_val = _cv(ws, row, COL_C)
            d_val = _cv(ws, row, COL_D)

            if "PROPRIETARY" in c_val.upper() and d_val.upper() == "LENOVO DEVELOPED":
                for col in (COL_I, COL_J, COL_K, COL_L):
                    ws.cell(row=row, column=col).value = "N/A"
                filled += 1
                continue

            lc = _classify_license(c_val)
            if lc == "GPL":
                ws.cell(row=row, column=COL_I).value = "Yes"
                ws.cell(row=row, column=COL_J).value = "Yes"
                ws.cell(row=row, column=COL_K).value = "Yes"
                ws.cell(row=row, column=COL_L).value = "Required"
                filled += 1
            elif lc == "PERMISSIVE":
                for col, val in [(COL_I, "NO"), (COL_J, "NO"), (COL_K, "YES"), (COL_L, "Not required")]:
                    if not _cv(ws, row, col):
                        ws.cell(row=row, column=col).value = val
                filled += 1
            else:
                skipped += 1

        print(f"\n[Phase 4] I/J/K/L: Filled={filled}, Skipped(unknown/see-below)={skipped}")

    # ------------------------------------------------------------------
    # Phase 5 — npm Cross-reference
    # ------------------------------------------------------------------
    def _phase5_npm(self, ws, groups: dict, known_npm: set[str]) -> tuple[int, list]:
        yellow = 0
        ok = 0
        unverified: list[tuple] = []

        for row, grp in groups.items():
            if grp not in ("NPM", "TRANSITIVE"):
                continue
            a = _cv(ws, row, COL_A)
            if a and a not in known_npm:
                ws.cell(row=row, column=COL_A).fill = YELLOW
                yellow += 1
                unverified.append((grp, row, a))
            else:
                ok += 1

        print(f"[Phase 5] npm/transitive: OK={ok}, TO BE VERIFIED={yellow}")
        return yellow, unverified

    # ------------------------------------------------------------------
    # Phase 6 — License Consistency (Col C fills)
    # ------------------------------------------------------------------
    def _phase6_license_c(self, ws, groups: dict) -> None:
        gray = white = 0
        for row, grp in groups.items():
            if grp in ("SKIP", "IGNORE"):
                continue
            c_val = _cv(ws, row, COL_C)
            if c_val.lower() == "see below":
                ws.cell(row=row, column=COL_C).fill = GRAY15
                gray += 1
            elif not c_val:
                ws.cell(row=row, column=COL_C).fill = WHITE
                white += 1
        print(f"\n[Phase 6] Col C: Gray('see below')={gray}, White(empty cleared)={white}")

    # ------------------------------------------------------------------
    # Phase 7 — Col F/G Validation
    # ------------------------------------------------------------------
    def _phase7_fg_validate(self, ws, groups: dict) -> tuple[int, int]:
        f_pink = g_pink = 0
        for row, grp in groups.items():
            if grp in ("SKIP", "IGNORE"):
                continue
            # Skip rows where Col C = "IGNORE"
            if _cv(ws, row, COL_C).upper() == "IGNORE":
                continue
            a = _cv(ws, row, COL_A)
            if not EXT_RE.search(a):
                continue
            if not _cv(ws, row, COL_F):
                ws.cell(row=row, column=COL_F).fill = PINK
                f_pink += 1
            if not _cv(ws, row, COL_G):
                ws.cell(row=row, column=COL_G).fill = PINK
                g_pink += 1
        print(f"\n[Phase 7] Col F/G: F_pink={f_pink}, G_pink={g_pink}")
        return f_pink, g_pink

    # ------------------------------------------------------------------
    # Phase 8 — OneCLI SBOM cross-reference
    # ------------------------------------------------------------------
    def _phase8_onecli_xref(self, ws, groups: dict) -> tuple[int, int, list]:
        blue = orange = 0
        orange_list: list[tuple] = []

        if not Path(self.onecli_json_path).exists():
            print(f"\n[Phase 8] PENDING — not found: {self.onecli_json_path}")
            return 0, 0, []

        with open(self.onecli_json_path) as f:
            onecli_names: set[str] = set(json.load(f).keys())

        print(f"\n[Phase 8] OneCLI SBOM JSON ({self.cfg.name}): {len(onecli_names)} output entries")

        for row, grp in groups.items():
            if grp != "INPUT_ONECLI":
                continue
            a = _cv(ws, row, COL_A)
            if a in onecli_names:
                ws.cell(row=row, column=COL_A).fill = BLUE
                blue += 1
            else:
                ws.cell(row=row, column=COL_A).fill = ORANGE
                orange += 1
                orange_list.append((row, a))

        print(f"[Phase 8] Input(OneCLI) vs {self.cfg.name} OneCLI Output: BLUE={blue}, ORANGE={orange}")
        if orange_list:
            for r, name in orange_list[:20]:
                print(f"           Row {r}: {name!r}")
        return blue, orange, orange_list

    # ------------------------------------------------------------------
    # Manual actions detection
    # ------------------------------------------------------------------
    def _detect_manual_actions(self, orange_list: list[str]) -> list[str]:
        actions = []
        if "icudtl.dat" in orange_list:
            actions.append(
                "ACTION REQUIRED: Add 'icudtl.dat' row to Input (Electron) section manually "
                f"(Col E: {self.cfg.input_electron_marker}icudtl.dat)"
            )
        return actions

    # ------------------------------------------------------------------
    # Summary helpers
    # ------------------------------------------------------------------
    def _build_summary(self, ws, groups: dict) -> dict:
        row_sets = {
            "Output":     {r for r, g in groups.items() if g == "OUTPUT"},
            "Input":      {r for r, g in groups.items() if g in ("INPUT_SOURCE","INPUT_ONECLI","INPUT_ELECTRON")},
            "npm":        {r for r, g in groups.items() if g == "NPM"},
            "Transitive": {r for r, g in groups.items() if g == "TRANSITIVE"},
        }
        summary = {}
        for label, row_set in row_sets.items():
            summary[label] = {
                "BLUE":    _fill_count(ws, row_set, COL_A, "ADD8E6"),
                "ORANGE":  _fill_count(ws, row_set, COL_A, "FFA500"),
                "PINK_A":  _fill_count(ws, row_set, COL_A, "FFB6C1"),
                "YEL_A":   _fill_count(ws, row_set, COL_A, "FFFF00"),
                "GRAY_C":  _fill_count(ws, row_set, COL_C, "D9D9D9"),
            }
        return summary

    # ------------------------------------------------------------------
    # Print report
    # ------------------------------------------------------------------
    def _print_report(self, r: ReviewResult) -> None:
        print("\n" + "=" * 70)
        print(f"  SBOM REVIEW SUMMARY — {r.platform.upper()}")
        print("=" * 70)

        hdr = f"{'Group':<16} | {'BLUE':>5} | {'ORANGE':>6} | {'PINK(A)':>7} | {'YEL(A)':>6} | {'GRAY(C)':>7}"
        print(hdr)
        print("-" * len(hdr))
        for label in ("Output", "Input", "npm", "Transitive"):
            s = r.summary_rows.get(label, {})
            print(f"{label:<16} | {s.get('BLUE',0):>5} | {s.get('ORANGE',0):>6} | "
                  f"{s.get('PINK_A',0):>7} | {s.get('YEL_A',0):>6} | {s.get('GRAY_C',0):>7}")

        cfg = self.cfg
        print(f"\n[Phase 8] Input(OneCLI) BLUE={r.phase8_blue}, ORANGE={r.phase8_orange}")

        print("\n=== ORANGE (MISMATCH) — Phase 2 Output vs Input ===")
        unique_orange = sorted(set(r.orange_list))
        for name in unique_orange:
            known = name in cfg.known_orange_names
            is_dist = cfg.dist_pkg_predicate and cfg.dist_pkg_predicate(name)
            if known or is_dist:
                tag = "[known exception]"
            else:
                tag = "[UNEXPECTED — review]"
            print(f"  {tag}  {name!r}")

        print(f"\n=== PINK (NOT DISTRIBUTED) — {len(set(r.pink_list))} items ===")
        for name in sorted(set(r.pink_list)):
            print(f"  {name!r}")

        if r.phase5_unverified:
            print(f"\n=== TO BE VERIFIED (Yellow Col A) — {r.phase5_yellow} ===")
            for grp, row, name in sorted(r.phase5_unverified):
                print(f"  [{grp}] Row {row}: {name!r}")

        if r.phase7_f_pink or r.phase7_g_pink:
            print(f"\n=== Col F/G MISSING — F_pink={r.phase7_f_pink}, G_pink={r.phase7_g_pink} ===")

        if r.manual_actions:
            print("\n=== ⚠ MANUAL ACTIONS REQUIRED ===")
            for action in r.manual_actions:
                print(f"  !! {action}")

        print(f"\n[Done] Saved → {r.output_path}")


# ---------------------------------------------------------------------------
# Convenience factory
# ---------------------------------------------------------------------------
def review_xlsx(
    sbom_path: str,
    platform: Optional[str] = None,
    deps_info_path: Optional[str] = None,
    lock_path: Optional[str] = None,
    fossa_json_path: Optional[str] = None,
    onecli_json_path: Optional[str] = None,
    output_path: Optional[str] = None,
) -> ReviewResult:
    reviewer = XlsxReviewer(
        sbom_path=sbom_path,
        platform=platform,
        deps_info_path=deps_info_path,
        lock_path=lock_path,
        fossa_json_path=fossa_json_path,
        onecli_json_path=onecli_json_path,
        output_path=output_path,
    )
    return reviewer.run()
