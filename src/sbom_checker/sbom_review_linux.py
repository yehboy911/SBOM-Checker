#!/usr/bin/env python3
"""
SBOM Excel Review Script — UpdateXpress v5.4.0 (Linux / tgz)
Methodology: OSC-compliance-sbom-excel.md

Phases:
  1  Group Classification
  2  Bidirectional Matching (Output ↔ Input)
  3  Input (OneCLI) Col C → "see below" + Gray 15%
  4  Col I/J/K/L Fill Rules
  5  npm Cross-reference (Col A yellow)
  6  License Consistency (Col C: "see below"→Gray, empty→White/clear)
  7  Col F/G Validation (file-extension rows must have F and G filled)
  8  OneCLI SBOM Cross-reference — Linux onecli_linux_sbom_data.json
  9  Save & Summary
"""
import json
import re
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import PatternFill

SBOM_PATH = (
    "/Users/Yehboy/Library/CloudStorage/OneDrive-Lenovo/文件/OSC/LXCE/"
    "UpdateXpress/LXCE_UpdateXpress_v5.4.0/Dev/"
    "updatexpress (tgz), v5.4.0 (SI Agile 25-2)-SBOM-30Sep2025-DRAFT.xlsx"
)
DEPS_INFO_PATH = "/Users/Yehboy/OSC/LXCE_5.4.0/UpdateXpress/dependencies_info.json"
LOCK_PATH      = "/Users/Yehboy/OSC/LXCE_5.4.0/UpdateXpress/package-lock.json"
LIN_JSON       = "/Users/Yehboy/OSC/LXCE_5.4.0/UpdateXpress/docs/onecli_linux_sbom_data.json"
OUTPUT_PATH    = SBOM_PATH.replace("-DRAFT.xlsx", "-DRAFT_reviewed.xlsx")

# --- Colors (PatternFill, NOT ConditionalFormatting) ---
BLUE   = PatternFill("solid", start_color="ADD8E6", end_color="ADD8E6")
PINK   = PatternFill("solid", start_color="FFB6C1", end_color="FFB6C1")
ORANGE = PatternFill("solid", start_color="FFA500", end_color="FFA500")
YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
GRAY15 = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
WHITE  = PatternFill("none")   # clear / no fill

# Columns (1-indexed in openpyxl)
COL_A, COL_B, COL_C, COL_D, COL_E, COL_F, COL_G, COL_H = 1, 2, 3, 4, 5, 6, 7, 8
COL_I, COL_J, COL_K, COL_L = 9, 10, 11, 12

DEEP_DEP_MARKER = "Deep dependencies"

# --- File-extension → Col F (USE) mapping (Linux) ---
EXT_RE = re.compile(r'\.[a-zA-Z0-9]{2,6}$')

def ext_to_use(filename):
    """Map filename to expected Col F (USE column) value — Linux variant."""
    m = EXT_RE.search(filename.lower())
    if not m:
        return None
    ext = m.group()
    MAP = {
        '.dll': 'DLL', '.exe': 'EXE', '.pak': 'PAK', '.zip': 'ZIP',
        '.asar': 'ASAR', '.dat': 'File', '.sys': 'File', '.cat': 'File',
        '.txt': 'File', '.pdf': 'File', '.png': 'File', '.gif': 'File',
        '.so':  'SO',   # Linux shared object (equivalent to .dll)
        '.tgz': 'TGZ',
        '.rpm': 'RPM',
    }
    if ext == '.bin':
        name_lower = filename.lower()
        if 'snapshot' in name_lower or 'blob' in name_lower:
            return 'Binary'    # V8 snapshot/blob files
        return 'BIN'           # firmware / Linux binary
    return MAP.get(ext, 'File')

def cv(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

# ============================================================
# Phase 5 data — build known_npm_names
# Priority: dependencies_info.json (primary) > package-lock.json (fallback)
# ============================================================
known_npm_names = set()

try:
    with open(DEPS_INFO_PATH) as f:
        dinfo = json.load(f)
    known_npm_names.update(dinfo.keys())
    print(f"[Phase 5] dependencies_info.json: {len(dinfo)} direct npm names loaded")
except Exception as e:
    print(f"[Phase 5] WARN: could not load dependencies_info.json: {e}")

lock_names = set()
try:
    with open(LOCK_PATH) as f:
        lock = json.load(f)
    if "packages" in lock:
        for pkg_path in lock["packages"]:
            if pkg_path and "node_modules/" in pkg_path:
                lock_names.add(pkg_path.split("node_modules/")[-1])
    elif "dependencies" in lock:
        lock_names.update(lock["dependencies"].keys())
    new_from_lock = lock_names - known_npm_names
    known_npm_names.update(new_from_lock)
    print(f"[Phase 5] package-lock.json fallback: +{len(new_from_lock)} transitive names")
    print(f"[Phase 5] Total known npm names: {len(known_npm_names)}")
except Exception as e:
    print(f"[Phase 5] WARN: could not load package-lock.json: {e}")

# ============================================================
# Load workbook
# ============================================================
wb = openpyxl.load_workbook(SBOM_PATH)
ws = wb.active
total_rows = ws.max_row
print(f"\n[Load] {total_rows} rows, {ws.max_column} columns")

# ============================================================
# Phase 1 — Group Classification (Linux paths: forward slash)
# ============================================================
groups = {}
deep_dep_row = None

for row in range(1, total_rows + 1):
    a = cv(ws, row, COL_A)
    c = cv(ws, row, COL_C)
    e = cv(ws, row, COL_E)
    f = cv(ws, row, COL_F)

    if row <= 20:
        groups[row] = "SKIP"
        continue
    if a.startswith("*BUILD OUTPUT"):
        groups[row] = "SKIP"
        continue
    if a in ("Direct dependencies",):
        groups[row] = "SKIP"
        continue
    if a == DEEP_DEP_MARKER:
        groups[row] = "SKIP"
        deep_dep_row = row
        continue
    if not a:
        groups[row] = "SKIP"
        continue

    if deep_dep_row and row > deep_dep_row:
        groups[row] = "TRANSITIVE"
        continue

    # Linux output target path (forward slash)
    if "Target/lnvgy_utl_lxce_ux_linux_x86-64/" in e:
        groups[row] = "OUTPUT"
        continue
    if e.startswith("Target/") and "lnvgy_utl_lxce_ux" in e:
        groups[row] = "OUTPUT"
        continue

    # Input source (FOSSA scan paths)
    if e.startswith("1/"):
        groups[row] = "INPUT_SOURCE"
        continue

    # Input OneCLI (Linux forward slash)
    if "updatexpress/bin/command/onecli/" in e:
        groups[row] = "INPUT_ONECLI"
        continue

    # Input Electron (Linux forward slash)
    if "updatexpress/bin/electron/" in e:
        groups[row] = "INPUT_ELECTRON"
        continue

    if f.lower() == "npm" or f.lower().startswith("npm"):
        groups[row] = "NPM"
        continue
    if c == "NOT DISTRIBUTED":
        groups[row] = "SKIP"
        continue

    groups[row] = "SKIP"

group_counts = Counter(groups.values())
print(f"[Phase 1] Groups: {dict(group_counts)}")
if deep_dep_row:
    print(f"         Deep dependencies separator at row {deep_dep_row}")

# ============================================================
# Phase 2 — Bidirectional Matching (Output ↔ UX Input)
# ============================================================
output_names = defaultdict(list)
input_names  = defaultdict(list)

for row, grp in groups.items():
    a = cv(ws, row, COL_A)
    if grp == "OUTPUT":
        output_names[a].append(row)
    elif grp in ("INPUT_SOURCE", "INPUT_ONECLI", "INPUT_ELECTRON"):
        input_names[a].append(row)

print(f"\n[Phase 2] output_names: {len(output_names)} unique | input_names: {len(input_names)} unique")

stats = {"BLUE_out": 0, "ORANGE_out": 0, "BLUE_in": 0, "PINK_in": 0, "SKIP_in": 0}
orange_list = []
pink_list   = []

for name, rows in output_names.items():
    if name in input_names:
        for row in rows:
            ws.cell(row=row, column=COL_A).fill = BLUE
        stats["BLUE_out"] += len(rows)
    else:
        for row in rows:
            ws.cell(row=row, column=COL_A).fill = ORANGE
        stats["ORANGE_out"] += len(rows)
        orange_list.append(name)

for name, rows in input_names.items():
    skip_all = all(
        "PROPRIETARY" in cv(ws, r, COL_C).upper()
        and cv(ws, r, COL_D).upper() == "LENOVO DEVELOPED"
        for r in rows
    )
    if skip_all:
        stats["SKIP_in"] += len(rows)
        continue

    if name in output_names:
        for row in rows:
            ws.cell(row=row, column=COL_A).fill = BLUE
        stats["BLUE_in"] += len(rows)
    else:
        for row in rows:
            ws.cell(row=row, column=COL_A).fill = PINK
            existing_c = cv(ws, row, COL_C)
            if not existing_c:
                ws.cell(row=row, column=COL_C).value = "NOT DISTRIBUTED"
        stats["PINK_in"] += len(rows)
        pink_list.append(name)

print(f"[Phase 2] Output: BLUE={stats['BLUE_out']}, ORANGE={stats['ORANGE_out']}")
print(f"[Phase 2] Input:  BLUE={stats['BLUE_in']}, PINK={stats['PINK_in']}, "
      f"Skipped(PROP+LENOVO)={stats['SKIP_in']}")

# ============================================================
# Phase 3 — Input (OneCLI) Col C → "see below" + GRAY15
# Actual match against OneCLI Output Col A is Phase 8
# ============================================================
p3_filled = 0
for row, grp in groups.items():
    if grp != "INPUT_ONECLI":
        continue
    if not cv(ws, row, COL_C):
        ws.cell(row=row, column=COL_C).value = "see below"
        ws.cell(row=row, column=COL_C).fill = GRAY15
        p3_filled += 1

print(f"\n[Phase 3] Input(OneCLI) Col C filled 'see below': {p3_filled} rows")

# ============================================================
# Phase 4 — Col I / J / K / L Fill Rules
# ============================================================
PERMISSIVE_TOKENS = {
    "mit", "isc", "bsd", "bsd-2-clause", "bsd-3-clause",
    "apache", "apache-2.0", "x11", "unicode", "zlib", "cc0",
    "cc0-1.0", "wtfpl", "unlicense", "0bsd", "ms-eula",
    "commercial", "proprietary", "public domain"
}

def classify_license(c_val):
    if not c_val or c_val.lower() in ("see below",):
        return "UNKNOWN"
    c_lower = c_val.lower()
    if "lgpl" in c_lower or ("gpl" in c_lower and "lgpl" not in c_lower):
        return "GPL"
    tokens = [t.strip().lower() for t in c_lower.replace(";", ",").split(",") if t.strip()]
    if tokens and all(any(p in t for p in PERMISSIVE_TOKENS) for t in tokens):
        return "PERMISSIVE"
    return "UNKNOWN"

phase4_filled = 0
phase4_skipped = 0
input_groups_set = {"INPUT_SOURCE", "INPUT_ONECLI", "INPUT_ELECTRON", "NPM", "TRANSITIVE"}

for row, grp in groups.items():
    if grp not in input_groups_set:
        continue

    c_val = cv(ws, row, COL_C)
    d_val = cv(ws, row, COL_D)

    if "PROPRIETARY" in c_val.upper() and d_val.upper() == "LENOVO DEVELOPED":
        ws.cell(row=row, column=COL_I).value = "N/A"
        ws.cell(row=row, column=COL_J).value = "N/A"
        ws.cell(row=row, column=COL_K).value = "N/A"
        ws.cell(row=row, column=COL_L).value = "N/A"
        phase4_filled += 1
        continue

    lc = classify_license(c_val)

    if lc == "GPL":
        ws.cell(row=row, column=COL_I).value = "Yes"
        ws.cell(row=row, column=COL_J).value = "Yes"
        ws.cell(row=row, column=COL_K).value = "Yes"
        ws.cell(row=row, column=COL_L).value = "Required"
        phase4_filled += 1

    elif lc == "PERMISSIVE":
        if not cv(ws, row, COL_I):
            ws.cell(row=row, column=COL_I).value = "NO"
        if not cv(ws, row, COL_J):
            ws.cell(row=row, column=COL_J).value = "NO"
        if not cv(ws, row, COL_K):
            ws.cell(row=row, column=COL_K).value = "YES"
        if not cv(ws, row, COL_L):
            ws.cell(row=row, column=COL_L).value = "Not required"
        phase4_filled += 1

    else:
        phase4_skipped += 1

print(f"\n[Phase 4] I/J/K/L: Filled={phase4_filled}, Skipped(unknown)={phase4_skipped}")

# ============================================================
# Phase 5 — npm Cross-reference (Col A YELLOW)
# ============================================================
phase5_yellow     = 0
phase5_ok         = 0
phase5_unverified = []

for row, grp in groups.items():
    if grp not in ("NPM", "TRANSITIVE"):
        continue
    a = cv(ws, row, COL_A)
    if a and a not in known_npm_names:
        ws.cell(row=row, column=COL_A).fill = YELLOW
        phase5_yellow += 1
        phase5_unverified.append((grp, row, a))
    else:
        phase5_ok += 1

print(f"[Phase 5] npm/transitive: OK={phase5_ok}, TO BE VERIFIED={phase5_yellow}")

# ============================================================
# Phase 6 — License Consistency (Col C fills)
# "see below" → GRAY 15%
# empty       → WHITE (clear fill)
# ============================================================
phase6_gray  = 0
phase6_white = 0

for row, grp in groups.items():
    if grp == "SKIP":
        continue
    c_val = cv(ws, row, COL_C)
    if c_val.lower() == "see below":
        ws.cell(row=row, column=COL_C).fill = GRAY15
        phase6_gray += 1
    elif not c_val:
        ws.cell(row=row, column=COL_C).fill = WHITE
        phase6_white += 1

print(f"\n[Phase 6] Col C: Gray('see below')={phase6_gray}, White(empty cleared)={phase6_white}")

# ============================================================
# Phase 7 — Col F/G Validation
# ============================================================
p7_f_pink = 0
p7_g_pink = 0

for row, grp in groups.items():
    if grp == "SKIP":
        continue
    a = cv(ws, row, COL_A)
    if not EXT_RE.search(a):
        continue

    f_val = cv(ws, row, COL_F)
    g_val = cv(ws, row, COL_G)

    if not f_val:
        ws.cell(row=row, column=COL_F).fill = PINK
        p7_f_pink += 1

    if not g_val:
        ws.cell(row=row, column=COL_G).fill = PINK
        p7_g_pink += 1

print(f"\n[Phase 7] Col F/G: F_pink={p7_f_pink}, G_pink={p7_g_pink}")

# ============================================================
# Phase 8 — OneCLI SBOM Cross-reference (Linux)
# Match INPUT_ONECLI Col A vs onecli_linux_sbom_data.json Output keys
# ============================================================
p8_blue = p8_orange = 0
p8_orange_list = []

try:
    with open(LIN_JSON) as f:
        lin_names = set(json.load(f).keys())
    print(f"\n[Phase 8] onecli_linux_sbom_data.json: {len(lin_names)} Output entries")

    for row, grp in groups.items():
        if grp != "INPUT_ONECLI":
            continue
        a = cv(ws, row, COL_A)
        if a in lin_names:
            ws.cell(row=row, column=COL_A).fill = BLUE
            p8_blue += 1
        else:
            ws.cell(row=row, column=COL_A).fill = ORANGE
            p8_orange += 1
            p8_orange_list.append((row, a))

    print(f"[Phase 8] Input(OneCLI) vs Linux OneCLI Output: BLUE={p8_blue}, ORANGE={p8_orange}")
    if p8_orange_list:
        print("         ORANGE (not in OneCLI Linux Output):")
        for r, name in p8_orange_list[:30]:
            print(f"           Row {r}: {name!r}")

except FileNotFoundError:
    print("\n[Phase 8] PENDING — onecli_linux_sbom_data.json not found")

# ============================================================
# Phase 9 — Save & Summary
# ============================================================
wb.save(OUTPUT_PATH)
print(f"\n[Phase 9] Saved → {OUTPUT_PATH}")

wb2 = openpyxl.load_workbook(OUTPUT_PATH)
ws2 = wb2.active

def cnt(row_set, col, hex_color):
    c = 0
    for row in row_set:
        cell = ws2.cell(row=row, column=col)
        if cell.fill and cell.fill.fill_type == "solid":
            if hex_color.upper() in cell.fill.fgColor.rgb.upper():
                c += 1
    return c

rows_out   = {r for r, g in groups.items() if g == "OUTPUT"}
rows_in    = {r for r, g in groups.items() if g in ("INPUT_SOURCE","INPUT_ONECLI","INPUT_ELECTRON")}
rows_npm   = {r for r, g in groups.items() if g == "NPM"}
rows_trans = {r for r, g in groups.items() if g == "TRANSITIVE"}

print("\n=== SUMMARY TABLE ===")
hdr = f"{'Group':<16} | {'BLUE':>5} | {'ORANGE':>6} | {'PINK(A)':>7} | {'YEL(A)':>6} | {'GRAY(C)':>7}"
print(hdr)
print("-" * len(hdr))
for label, row_set in [("Output",rows_out),("Input",rows_in),("npm",rows_npm),("Transitive",rows_trans)]:
    b  = cnt(row_set, COL_A, "ADD8E6")
    o  = cnt(row_set, COL_A, "FFA500")
    p  = cnt(row_set, COL_A, "FFB6C1")
    ya = cnt(row_set, COL_A, "FFFF00")
    gc = cnt(row_set, COL_C, "D9D9D9")
    print(f"{label:<16} | {b:>5} | {o:>6} | {p:>7} | {ya:>6} | {gc:>7}")

print("\n=== ORANGE (MISMATCH) — Phase 2 Output vs Input ===")
for name in sorted(set(orange_list)):
    print(f"  {name!r}")

print(f"\n=== PINK (NOT DISTRIBUTED in Output) — {len(set(pink_list))} items ===")
for name in sorted(set(pink_list)):
    print(f"  {name!r}")

if phase5_unverified:
    print(f"\n=== TO BE VERIFIED (Yellow Col A) — {len(phase5_unverified)} ===")
    for grp, row, name in sorted(phase5_unverified):
        c = cv(ws2, row, COL_C)
        print(f"  [{grp}] Row {row}: {name!r}  License={c!r}")

print("\n=== Col F/G Extension Reference (Linux) ===")
print("  .so→SO | .dll→DLL | .exe→EXE | .pak→PAK | .zip→ZIP | .tgz→TGZ | .asar→ASAR")
print("  .bin→Binary(snapshot/blob) | .bin→BIN(firmware/Linux binary) | default→File")
print("  Col G default = 'AS IS'")

# Known Linux ORANGE exceptions (Phase 2 — not gaps, manual action only)
LINUX_ORANGE_EXCEPTIONS = {
    "lxce_ux.bin",       # Linux electron binary (renamed from 'electron')
    "app.asar",          # Electron ASAR bundle (built, no single input)
    "icudtl.dat",        # ICU data file — needs manual Input entry
}
# tgz distribution package (matches pattern)
def is_linux_tgz_exception(name):
    return name.startswith("lnvgy_utl_lxce_ux") and name.endswith("_linux_indiv.tgz")

phase2_orange = sorted(set(orange_list))
if phase2_orange:
    print("\n=== Phase 2 ORANGE — Known vs Unexpected ===")
    for name in phase2_orange:
        known = name in LINUX_ORANGE_EXCEPTIONS or is_linux_tgz_exception(name)
        tag = "[known exception]" if known else "[UNEXPECTED — review]"
        print(f"  {tag}  {name!r}")
